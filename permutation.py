'''
	Permutation test of dependency of two categorical variables.

	Created on 23. 1. 2018

	@author: Peter Demjan <peter.demjan@gmail.com>
	
	Required software and modules:
		Python 3.5 (https://www.python.org)
		openpyxl 2.4.7 (https://openpyxl.readthedocs.io)
		numpy 1.11.0 (http://www.numpy.org)
	
'''

from openpyxl import (load_workbook, Workbook)
from openpyxl.styles import colors
import numpy as np
import time, datetime

file_in = "data\\data_in.xlsx" # input data file
file_out = "result.xlsx" # output data file
rand_level = 90 # level of randomness, e.g. 90 means observed value is considered non-random, if it is higher than 90% randomly generated values
converg_diff = 0.05 # convergence threshold (e.g. 0.05 means difference between subsequent results must be <= 5%)
iters_start = 1000 # starting number of iterations when computing randomized values

def calc_dependence(cat_a, cat_b, data, rev = 0):
	# cat_a: observed independent category
	# cat_b: observed dependent category
	# data: [[cat_1, cat_2], ...]; cat_1: column with independent categories, cat_2: column with dependent categories
	# rev: if 1, reverse independent and dependent category
	
	if rev:
		cat_1, cat_2 = cat_b, cat_a
	else:
		cat_1, cat_2 = cat_a, cat_b
	
	slice_1 = data[data[:,1 if rev else 0] == cat_1]
	slice_2 = slice_1[slice_1[:,0 if rev else 1] == cat_2]
	return slice_2.shape[0] / slice_1.shape[0]


# add timestamp to output filename
# format will be e.g. result_[yyyymmdd]_[hhmmss].xlsx
file_out = ("%s_%s.%s" % (file_out.split(".")[0], datetime.datetime.fromtimestamp(time.time()).strftime('%Y%m%d_%H%M%S'), file_out.split(".")[1]))

ws = load_workbook(filename = file_in, read_only = True)
ws = ws[ws.get_sheet_names()[0]]

# load data from Excel file, first sheet
# first row is header
# column 1 is row ID
# columns 2 and 3 are the observed categories
# other columns and sheets are ignored
data = []
for row in ws.iter_rows(min_row = 2):
	data.append([(None if (cell.value is None) else str(cell.value).strip()) for cell in row[:3]])
ws = None
data = np.array(data, dtype = object)
data = data[:,1:]

cats_a = data[:,0]
cats_a = np.unique(cats_a[cats_a != None])
cats_b = data[:,1]
cats_b = np.unique(cats_b[cats_b != None])

collect = [] # [[cat_a, cat_b, r_obs, r_rnd], ...]
data_rnd_a, data_rnd_b = data.copy().T
cmax = 2 * cats_a.shape[0] * cats_b.shape[0]
c = 1
for rev in [0,1]:
	for cat_a in cats_a:
		for cat_b in cats_b:
			# for each combination of categories, calculate the ratio at which they occur together
			r_obs = calc_dependence(cat_a, cat_b, data, rev = rev)
			
			iters = iters_start
			r_last = None
			r_cnt = 0
			doubled = False
			while True:
				print("\r%s / %s%s (%d/%d) iters: %d           " % (cat_a, cat_b, " rev" if rev else "", c, cmax, iters), end = "")
				r_rnd = []
				for i in range(iters):
					np.random.shuffle(data_rnd_a)
					np.random.shuffle(data_rnd_b)
					r_rnd.append(calc_dependence(cat_a, cat_b, np.vstack((data_rnd_a, data_rnd_b)).T, rev = rev))
				r_rnd = np.array(r_rnd)
				r_rnd = np.percentile(r_rnd, rand_level)
				if not r_last is None:
					if abs(r_rnd - r_last) <= r_rnd * converg_diff:
						r_cnt += 1
						if r_cnt == 2:
							if doubled:
								break
							else:
								doubled = True
								r_cnt = 0
					else:
						r_cnt = 0
						doubled = False
				r_last = r_rnd
				if not r_cnt:
					iters *= 2
			
			if rev:
				collect.append([cat_b, cat_a, r_obs, r_rnd])
			else:
				collect.append([cat_a, cat_b, r_obs, r_rnd])
			
			c += 1

# save results

wb = Workbook()
ws = wb.active
ws.title = "Observed"
ws_rnd = wb.create_sheet(title = "Randomized")

cats = cats_a.tolist() + cats_b.tolist()
cell = ws.cell(column = 1, row = 1, value = "Indep.\\Dep.")
cell.font = cell.font.copy(bold = True)
cell = ws_rnd.cell(column = 1, row = 1, value = "Indep.\\Dep.")
cell.font = cell.font.copy(bold = True)

for i, cat in enumerate(cats):
	cell = ws.cell(column = i + 2, row = 1, value = cat)
	cell.font = cell.font.copy(bold = True)
	cell = ws.cell(column = 1, row = i + 2, value = cat)
	cell.font = cell.font.copy(bold = True)
	
	cell = ws_rnd.cell(column = i + 2, row = 1, value = cat)
	cell.font = cell.font.copy(bold = True)
	cell = ws_rnd.cell(column = 1, row = i + 2, value = cat)
	cell.font = cell.font.copy(bold = True)

for cat_a, cat_b, r_obs, r_rnd in collect:
	row = cats.index(cat_a) + 2
	col = cats.index(cat_b) + 2
	cell = ws.cell(column = col, row = row, value = r_obs)
	if r_obs > r_rnd:
		cell.font = cell.font.copy(bold = True, color = colors.RED)
	cell = ws_rnd.cell(column = col, row = row, value = r_rnd)

ws.cell(column = 1, row = len(cats) + 3, value = "Values in red are higher than %d%% of randomized results." % (rand_level))

wb.save(filename = file_out)
